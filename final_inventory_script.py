import openpyxl

class InventoryMachine:
    def __init__(self):
        self.file = 'inventory-file.xlsx'

    def get_sheet(self):
        try:
            doc = openpyxl.load_workbook(self.file, data_only=True)
            worksheet = doc.worksheets[-1]
            return worksheet
        except:
            return None

    def parse_column_row(self):
        sheet = self.get_sheet()
        data_start_row = 1
        sheet_title_num = 1
        data_start_col = 2
        data_end_num = sheet.max_row +1
        data_length = sheet.max_column + 1

        doc_range = range(data_start_row, data_length)
        length_range = range(data_start_col, data_end_num)
        if not sheet:
            return None
        data = [
            {
                sheet.cell(row=sheet_title_num, column=y)
                .value.lower()
                .replace(" ", "_"): sheet.cell(row=x, column=y)
                .value
                for y in doc_range
                if sheet.cell(row=x, column=y).value
                if sheet.cell(row=sheet_title_num, column=y).value
            }
            for x in length_range
        ]
        remove_empty_data = [x for x in data if bool(x)]
        return remove_empty_data

    def present_record(self):
        string = ''
        for index, data in enumerate(self.parse_column_row()):
            data.pop('record_code')
            data.pop('part_number')
            data.pop('part_description')
            data.pop('inventory_balance')
            string += 'Record {} \n'.format(index+1)
            for item, value in data.items():
                string += '{}: {} \n'.format(item.replace('_', ' ').capitalize(), value)
            string += 'Final inventory amount: {}'.format(
                data.get('open_inventory_amount')+data.get('amount_purchased')-data.get('amount_sold'))
            string += '\n\n'
        return string

    def present_data(self):
        print('All Inventory Record \n\n{}'.format(self.present_record()))
        print('Total Inventory: {}'.format(len(self.parse_column_row())))

inventory = InventoryMachine()
inventory.present_data()