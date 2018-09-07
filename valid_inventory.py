import openpyxl

class InventoryMachine:
    def __init__(self):
        self.file = 'inventory-file.xlsx'

    def get_sheet(self):
        try:
            doc = openpyxl.load_workbook(self.file, data_only=True)
            worksheet = doc.worksheets[-1]
            return worksheet
        except:
            return None

    def parse_column_row(self):
        sheet = self.get_sheet()
        data_start_row = 1
        sheet_title_num = 1
        data_start_col = 2
        data_end_num = sheet.max_row +1
        data_length = sheet.max_column + 1

        doc_range = range(data_start_row, data_length)
        length_range = range(data_start_col, data_end_num)
        if not sheet:
            return None
        data = [
            {
                sheet.cell(row=sheet_title_num, column=y)
                .value.lower()
                .replace(" ", "_"): sheet.cell(row=x, column=y)
                .value
                for y in doc_range
                if sheet.cell(row=x, column=y).value
                if sheet.cell(row=sheet_title_num, column=y).value
            }
            for x in length_range
        ]
        remove_empty_data = [x for x in data if bool(x)]
        return remove_empty_data

    def check_within_range(self, data):
        if data.startswith('AA'):
            end_code = int(data.split('AA')[1])
            return end_code in range(3000, 4000)

    def validate_record(self):
        data = [x for x in self.parse_column_row() if x.get('record_code') == 11]
        return data

    def valid_within_range(self):
        data = [x for x in self.validate_record() if self.check_within_range(x.get('part_number'))]
        return data

    def present_valid_record(self):
        string = ''
        for index, data in enumerate(self.validate_record()):
            data.pop('open_inventory_amount')
            data.pop('item_number')
            data.pop('amount_sold')
            data.pop('amount_purchased')
            string += 'Record {} \n'.format(index+1)
            for item, value in data.items():
                string += '{}: {} \n'.format(item.replace('_', ' ').capitalize(), value)
            string += '\n\n'
        return string

    def present_data(self):
        print('All Valid Inventory Record \n\n{}'.format(self.present_valid_record()))
        print('Total Inventory: {}'.format(len(self.parse_column_row())))
        print('Total Valid Record: {}'.format(len(self.validate_record())))
        print('Total Valid Record Within 3000 to 3999: {}'.format(len(self.valid_within_range())))

inventory = InventoryMachine()
inventory.present_data()